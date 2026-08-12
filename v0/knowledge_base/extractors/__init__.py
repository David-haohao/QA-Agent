# ============================================================
# 文档提取器 — PDF/Word/Excel/文本提取
# 支持数千个中文PDF文件的批量解析
# ============================================================

import os
import sys
import hashlib
import json
import subprocess
from typing import List, Dict, Optional

# 提高递归深度限制，防止复杂PDF导致RecursionError
sys.setrecursionlimit(10000)


class DocumentExtractor:
    """
    统一的文档内容提取器
    支持PDF(pdfplumber + PyPDF2)、Word、Excel和纯文本文件
    """

    def __init__(self, documents_dir: str, ocr_cache_dir: Optional[str] = None):
        """
        初始化提取器
        documents_dir: 存放PDF等文档的目录路径
        """
        self.documents_dir = documents_dir
        self.ocr_cache_dir = ocr_cache_dir or os.path.join(
            os.path.dirname(os.path.abspath(documents_dir)), "kb_data", "ocr_cache"
        )

    def extract_all(self) -> List[Dict]:
        """
        批量提取目录下所有文档的内容
        返回: [{"file_path": ..., "file_name": ..., "content": ..., "pages": [...]}, ...]
        """
        results, _ = self.extract_all_with_report()
        return results

    def extract_all_with_report(self):
        """提取所有支持的文件，并显式记录未能提取的文件。"""
        results = []
        skipped_files = []
        scanned_files = 0
        for root, dirs, files in os.walk(self.documents_dir):
            for fname in files:
                file_path = os.path.join(root, fname)
                fname_lower = fname.lower()
                if fname_lower.endswith(".pdf"):
                    scanned_files += 1
                    extracted = self._extract_pdf(file_path, fname)
                elif fname_lower.endswith(".docx"):
                    scanned_files += 1
                    extracted = self._extract_docx(file_path, fname)
                elif fname_lower.endswith((".xlsx", ".xls")):
                    scanned_files += 1
                    extracted = self._extract_excel(file_path, fname)
                elif fname_lower.endswith(".doc"):
                    scanned_files += 1
                    extracted = self._extract_doc(file_path, fname)
                elif fname_lower.endswith((".txt", ".md", ".csv")):
                    scanned_files += 1
                    extracted = self._extract_text_file(file_path, fname)
                else:
                    continue  # 跳过不支持的文件类型
                if extracted and extracted.get("content"):
                    results.append(extracted)
                else:
                    skipped_files.append({
                        "file_name": fname,
                        "reason": "empty_or_unreadable",
                    })
        return results, {
            "scanned_files": scanned_files,
            "success_files": len(results),
            "skipped_files": skipped_files,
        }

    def _extract_pdf(self, file_path: str, file_name: str) -> Optional[Dict]:
        """解析PDF文件，提取文本和分页信息"""
        full_text = ""
        pages = []
        try:
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    page_text = page.extract_text()
                    if page_text:
                        pages.append({
                            "page_num": page_num + 1,
                            "text": page_text,
                        })
                        full_text += page_text + "\n"
        except Exception:
            pass

        if not full_text.strip():
            # pdfplumber无法提取文本时回退到PyPDF2。
            try:
                from PyPDF2 import PdfReader
                full_text = ""
                reader = PdfReader(file_path)
                for page in reader.pages:
                    text = page.extract_text()
                    if text:
                        full_text += text + "\n"
            except Exception:
                pass

        if not full_text.strip():
            ocr_pages = self._extract_pdf_with_ocr(file_path)
            if ocr_pages:
                pages = [
                    {"page_num": page_num, "text": page_text}
                    for page_num, page_text in ocr_pages
                ]
                full_text = "".join(f"{page_text}\n" for _, page_text in ocr_pages)

        if full_text.strip():
            return {
                "file_path": file_path,
                "file_name": file_name,
                "content": full_text,
                "pages": pages,
                "doc_id": hashlib.md5(file_path.encode()).hexdigest()[:12],
            }
        return None

    def _extract_pdf_with_ocr(self, file_path: str):
        """对没有文本层的扫描 PDF 执行本地中文 OCR。

        每份 PDF 在独立子进程中执行，避免 OCR 运行时的图像内存跨文件累积；
        成功结果缓存到本地，重建中断后可直接续用。
        """
        cache_path = self._ocr_cache_path(file_path)
        cached = self._read_ocr_cache(cache_path)
        if cached is not None and self._ocr_cache_completed(cache_path):
            return cached

        try:
            command = [
                sys.executable,
                "-m",
                "knowledge_base.extractors.ocr_worker",
                file_path,
                cache_path,
            ]
            result = subprocess.run(
                command,
                capture_output=True,
                text=True,
                encoding="utf-8",
                timeout=3600,
            )
            pages = self._read_ocr_cache(cache_path) or []
            if result.returncode != 0:
                raise RuntimeError(result.stderr.strip() or "OCR worker exited unsuccessfully")
            return pages
        except Exception as exc:
            print(f"OCR解析失败: {file_path}, 错误: {exc}")
            return []

    def _ocr_cache_path(self, file_path: str) -> str:
        """返回与文件内容绑定的 OCR 缓存路径。"""
        hasher = hashlib.sha256()
        with open(file_path, "rb") as source:
            for block in iter(lambda: source.read(1024 * 1024), b""):
                hasher.update(block)
        return os.path.join(self.ocr_cache_dir, f"{hasher.hexdigest()}.json")

    @staticmethod
    def _read_ocr_cache(cache_path: str):
        try:
            with open(cache_path, "r", encoding="utf-8") as cache_file:
                payload = json.load(cache_file)
            return [
                (int(page_num), str(page_text))
                for page_num, page_text in payload.get("pages", [])
                if str(page_text).strip()
            ]
        except FileNotFoundError:
            return None
        except (OSError, ValueError, TypeError):
            return None

    @staticmethod
    def _ocr_cache_completed(cache_path: str) -> bool:
        try:
            with open(cache_path, "r", encoding="utf-8") as cache_file:
                return bool(json.load(cache_file).get("completed", False))
        except (OSError, ValueError, TypeError):
            return False

    @staticmethod
    def _write_ocr_cache(cache_path: str, pages) -> None:
        os.makedirs(os.path.dirname(cache_path), exist_ok=True)
        temp_path = cache_path + ".tmp"
        with open(temp_path, "w", encoding="utf-8") as cache_file:
            json.dump({"pages": pages}, cache_file, ensure_ascii=False)
        os.replace(temp_path, cache_path)

    def _extract_docx(self, file_path: str, file_name: str) -> Optional[Dict]:
        """解析Word文档"""
        try:
            from docx import Document
            doc = Document(file_path)
            full_text = "\n".join([para.text for para in doc.paragraphs if para.text.strip()])
            if full_text.strip():
                return {
                    "file_path": file_path,
                    "file_name": file_name,
                    "content": full_text,
                    "pages": [],
                    "doc_id": hashlib.md5(file_path.encode()).hexdigest()[:12],
                }
        except Exception as e:
            print(f"无法解析Word文件: {file_path}, 错误: {e}")
        return None

    def _extract_doc(self, file_path: str, file_name: str) -> Optional[Dict]:
        """解析旧版 .doc 文件（Microsoft Word 97-2003 二进制格式）
        使用 textutil (macOS)、antiword/catdoc (Linux) 或 olefile 提取文本"""
        full_text = ""

        # 策略1: Windows Microsoft Word COM（支持中文旧版 .doc）
        full_text = self._extract_doc_with_word_com(file_path)

        # 策略2: macOS textutil（系统自带，最可靠）
        if not full_text:
            try:
                import subprocess
                result = subprocess.run(
                    ["textutil", "-convert", "txt", "-stdout", file_path],
                    capture_output=True, text=True, timeout=60,
                )
                if result.returncode == 0 and result.stdout.strip():
                    full_text = result.stdout.strip()
            except Exception:
                pass

        # 策略3: antiword (Linux 常见)
        if not full_text:
            try:
                import subprocess
                result = subprocess.run(
                    ["antiword", file_path],
                    capture_output=True, text=True, timeout=30,
                )
                if result.returncode == 0 and result.stdout.strip():
                    full_text = result.stdout.strip()
            except Exception:
                pass

        # 策略4: catdoc (Linux 备选)
        if not full_text:
            try:
                import subprocess
                result = subprocess.run(
                    ["catdoc", file_path],
                    capture_output=True, text=True, timeout=30,
                )
                if result.returncode == 0 and result.stdout.strip():
                    full_text = result.stdout.strip()
            except Exception:
                pass

        # 策略5: olefile 直接读取（最后手段）
        if not full_text:
            try:
                import olefile
                ole = olefile.OleFileIO(file_path)
                if ole.exists("WordDocument"):
                    stream = ole.openstream("WordDocument")
                    data = stream.read()
                    # 简单提取可读文本（Latin-1可打印字符）
                    text = "".join(chr(b) if 32 <= b < 127 or b in (10, 13) else " " for b in data)
                    # 压缩空白
                    import re
                    text = re.sub(r'\s{3,}', '\n', text)
                    if len(text.strip()) > 100:
                        full_text = text.strip()
            except Exception:
                pass

        if full_text and len(full_text) > 50:
            return {
                "file_path": file_path,
                "file_name": file_name,
                "content": full_text,
                "pages": [],
                "doc_id": hashlib.md5(file_path.encode()).hexdigest()[:12],
            }
        return None

    def _extract_doc_with_word_com(self, file_path: str) -> str:
        """使用本机 Microsoft Word 打开旧版 DOC 并取得正文文本。"""
        if os.name != "nt":
            return ""

        word = None
        document = None
        try:
            import win32com.client

            word = win32com.client.DispatchEx("Word.Application")
            word.Visible = False
            word.DisplayAlerts = 0
            document = word.Documents.Open(
                os.path.abspath(file_path),
                ReadOnly=True,
                AddToRecentFiles=False,
                Visible=False,
            )
            return (document.Content.Text or "").strip()
        except Exception as exc:
            print(f"Word COM解析失败: {file_path}, 错误: {exc}")
            return ""
        finally:
            if document is not None:
                try:
                    document.Close(False)
                except Exception:
                    pass
            if word is not None:
                try:
                    word.Quit()
                except Exception:
                    pass

    def _extract_excel(self, file_path: str, file_name: str) -> Optional[Dict]:
        """解析Excel文档（.xlsx用openpyxl, .xls用xlrd）"""
        fname_lower = file_name.lower()

        # .xls 旧格式 → 用xlrd
        if fname_lower.endswith(".xls") and not fname_lower.endswith(".xlsx"):
            return self._extract_xls(file_path, file_name)

        # .xlsx 新格式 → 用openpyxl
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) for cell in row if cell is not None])
                    if row_text.strip():
                        rows.append(row_text)
                if rows:
                    parts.append(f"[工作表: {sheet_name}]\n" + "\n".join(rows))
            full_text = "\n\n".join(parts)
            if full_text.strip():
                return {
                    "file_path": file_path,
                    "file_name": file_name,
                    "content": full_text,
                    "pages": [],
                    "doc_id": hashlib.md5(file_path.encode()).hexdigest()[:12],
                }
        except Exception as e:
            # openpyxl失败时尝试xlrd
            print(f"openpyxl解析失败，尝试xlrd: {file_path}")
            return self._extract_xls(file_path, file_name)
        return None

    def _extract_xls(self, file_path: str, file_name: str) -> Optional[Dict]:
        """使用xlrd解析旧版.xls文件（xlrd 2.0+已移除.xls支持，需安装xlrd<2.0）"""
        try:
            import xlrd
            # 检查xlrd版本，2.0+不支持.xls
            xlrd_version = tuple(int(x) for x in xlrd.__version__.split(".")[:2])
            if xlrd_version >= (2, 0):
                print(f"xlrd {xlrd.__version__} 不支持.xls格式: {file_path}")
                print("  请降级xlrd: pip install 'xlrd<2.0'")
                # 尝试用openpyxl兜底
                return self._extract_xlsx_with_openpyxl(file_path, file_name)

            wb = xlrd.open_workbook(file_path)
            parts = []
            for sheet_idx in range(wb.nsheets):
                sheet = wb.sheet_by_index(sheet_idx)
                sheet_name = sheet.name
                rows = []
                for row_idx in range(sheet.nrows):
                    row_values = sheet.row_values(row_idx)
                    row_text = " | ".join([str(cell) for cell in row_values if cell is not None and str(cell).strip()])
                    if row_text.strip():
                        rows.append(row_text)
                if rows:
                    parts.append(f"[工作表: {sheet_name}]\n" + "\n".join(rows))
            full_text = "\n\n".join(parts)
            if full_text.strip():
                return {
                    "file_path": file_path,
                    "file_name": file_name,
                    "content": full_text,
                    "pages": [],
                    "doc_id": hashlib.md5(file_path.encode()).hexdigest()[:12],
                }
        except ImportError:
            print(f"xlrd未安装，无法解析.xls文件: {file_path}")
            print("  请执行: pip install 'xlrd<2.0'")
        except Exception as e:
            print(f"无法解析.xls文件: {file_path}, 错误: {e}")
            print("  提示: 可将文件另存为.xlsx格式后重试")
        return None

    def _extract_xlsx_with_openpyxl(self, file_path: str, file_name: str) -> Optional[Dict]:
        """仅用openpyxl解析（兜底）"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) for cell in row if cell is not None])
                    if row_text.strip():
                        rows.append(row_text)
                if rows:
                    parts.append(f"[工作表: {sheet_name}]\n" + "\n".join(rows))
            full_text = "\n\n".join(parts)
            if full_text.strip():
                return {
                    "file_path": file_path,
                    "file_name": file_name,
                    "content": full_text,
                    "pages": [],
                    "doc_id": hashlib.md5(file_path.encode()).hexdigest()[:12],
                }
        except Exception as e:
            print(f"openpyxl也无法解析: {file_path}, 错误: {e}")
        return None

    def _extract_text_file(self, file_path: str, file_name: str) -> Optional[Dict]:
        """解析纯文本文件"""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()
            if content.strip():
                return {
                    "file_path": file_path,
                    "file_name": file_name,
                    "content": content,
                    "pages": [],
                    "doc_id": hashlib.md5(file_path.encode()).hexdigest()[:12],
                }
        except Exception as e:
            print(f"无法解析文本文件: {file_path}, 错误: {e}")
        return None
